import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ─── DADOS DE ENTRADA ────────────────────────────────────────────
df_raw = pd.read_excel(os.path.join(BASE_DIR, 'dados_conciliacao.xlsx'), header=0, index_col=0)
marketplaces = list(df_raw.columns)
qtd_repasses  = {m: float(df_raw.loc['quantidade mês', m]) for m in marketplaces}
tempo_repasse = {m: float(df_raw.loc['hora gasta para conclilar', m]) for m in marketplaces}

# ─── PREMISSAS ───────────────────────────────────────────────────
HORAS_MES_PESSOA  = 176
CUSTO_HORA        = 20.0
VARIACAO_TEMPO_LENTO  = 0.30   # +30% no tempo (cenário pessimista)
VARIACAO_TEMPO_RAPIDO = 0.30   # -30% no tempo (cenário otimista)
CUSTO_AUTOMACAO   = 15000.0

# ─── CÁLCULOS ────────────────────────────────────────────────────
total_repasses   = sum(qtd_repasses.values())
min_tempo        = min(tempo_repasse.values())
mais_rapido      = min(tempo_repasse, key=tempo_repasse.get)
horas_totais     = {m: qtd_repasses[m] * tempo_repasse[m] for m in marketplaces}
total_horas      = sum(horas_totais.values())
pessoas_parciais = {m: horas_totais[m] / HORAS_MES_PESSOA for m in marketplaces}
indice_complexidade = {m: tempo_repasse[m] / min_tempo for m in marketplaces}
pct_repasses     = {m: qtd_repasses[m] / total_repasses for m in marketplaces}
pct_horas        = {m: horas_totais[m] / total_horas for m in marketplaces}
delta_esforco    = {m: pct_horas[m] - pct_repasses[m] for m in marketplaces}
horas_ideais     = {m: qtd_repasses[m] * min_tempo for m in marketplaces}
horas_poupadas   = {m: horas_totais[m] - horas_ideais[m] for m in marketplaces}
pessoas_perdidas = {m: horas_poupadas[m] / HORAS_MES_PESSOA for m in marketplaces}
custo_mensal     = {m: horas_totais[m] * CUSTO_HORA for m in marketplaces}
payback_meses    = {m: CUSTO_AUTOMACAO / custo_mensal[m] if custo_mensal[m] > 0 else 999 for m in marketplaces}
score_ranking    = {m: horas_totais[m] * (tempo_repasse[m] * CUSTO_HORA) for m in marketplaces}
ranking          = sorted(marketplaces, key=lambda m: score_ranking[m], reverse=True)

# Sensibilidade de tempo — cenário lento (+%) e rápido (-%)
tempo_lento    = {m: tempo_repasse[m] * (1 + VARIACAO_TEMPO_LENTO)  for m in marketplaces}
tempo_rapido   = {m: tempo_repasse[m] * (1 - VARIACAO_TEMPO_RAPIDO) for m in marketplaces}
horas_lento    = {m: qtd_repasses[m] * tempo_lento[m]   for m in marketplaces}
horas_rapido   = {m: qtd_repasses[m] * tempo_rapido[m]  for m in marketplaces}
pessoas_lento  = {m: horas_lento[m]  / HORAS_MES_PESSOA for m in marketplaces}
pessoas_rapido = {m: horas_rapido[m] / HORAS_MES_PESSOA for m in marketplaces}
delta_h_lento  = {m: horas_lento[m]  - horas_totais[m]  for m in marketplaces}
delta_h_rapido = {m: horas_rapido[m] - horas_totais[m]  for m in marketplaces}
custo_lento    = {m: horas_lento[m]  * CUSTO_HORA for m in marketplaces}
custo_rapido   = {m: horas_rapido[m] * CUSTO_HORA for m in marketplaces}
delta_custo_lento  = {m: custo_lento[m]  - custo_mensal[m] for m in marketplaces}
delta_custo_rapido = {m: custo_rapido[m] - custo_mensal[m] for m in marketplaces}

# ─── ESTILOS ─────────────────────────────────────────────────────
C_DARK   = '1A3A5C'
C_MID    = '1F4E79'
C_LIGHT  = '2E75B6'
C_ALT    = 'D6E4F0'
C_WHITE  = 'FFFFFF'
C_WARN   = 'FFE699'
C_OK     = 'C6EFCE'
C_RED    = 'FFCCCC'
C_TOTAL  = 'BDD7EE'
C_CARD1  = '1F4E79'
C_CARD2  = '2E75B6'
C_CARD3  = '375623'
C_CARD4  = '833C00'
C_CARD5  = '4B0082'

THIN = Side(style='thin', color='B8CCE4')
BRD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def fnt(bold=False, sz=10, color=None, name='Arial'):
    return Font(name=name, bold=bold, size=sz, color=color or '000000')

def fill(color):
    return PatternFill('solid', start_color=color)

def cell_style(cell, value=None, bold=False, sz=10, fc=None, bg=None,
               align='center', border=True, color=None, wrap=False):
    if value is not None:
        cell.value = value
    cell.font = Font(name='Arial', bold=bold, size=sz, color=color or ('FFFFFF' if bg and bg != C_WHITE and bg != C_ALT and bg != C_TOTAL and bg != C_WARN and bg != C_OK and bg != C_RED else '000000'))
    if bg:
        cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if border:
        cell.border = BRD

def title_row(ws, text, ncols, row=1, bg=C_DARK, sz=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1)
    cell_style(c, text, bold=True, sz=sz, bg=bg, color=C_WHITE)
    ws.row_dimensions[row].height = 28

def hdr_row(ws, cols, row, bg=C_MID):
    for i, v in enumerate(cols, 1):
        cell_style(ws.cell(row, i), v, bold=True, bg=bg, color=C_WHITE, wrap=True)
    ws.row_dimensions[row].height = 30

def data_row(ws, row, values, alt=False, custom_bg=None):
    bg_default = C_ALT if alt else C_WHITE
    for i, v in enumerate(values, 1):
        bg = custom_bg.get(i, bg_default) if custom_bg else bg_default
        txt_color = '000000'
        c = ws.cell(row, i)
        c.value = v
        c.font = Font(name='Arial', size=10, color=txt_color)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal='left' if i == 1 else 'center', vertical='center')
        c.border = BRD

def total_row(ws, row, values):
    for i, v in enumerate(values, 1):
        c = ws.cell(row, i)
        c.value = v
        c.font = Font(name='Arial', bold=True, size=10)
        c.fill = fill(C_TOTAL)
        c.alignment = Alignment(horizontal='left' if i == 1 else 'center', vertical='center')
        c.border = BRD

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def kpi_card(ws, row, col, label, value, bg):
    """Draws a 3-row KPI card starting at (row, col) spanning 2 cols"""
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
    ws.merge_cells(start_row=row+2, start_column=col, end_row=row+2, end_column=col+1)
    top = ws.cell(row, col); top.fill = fill(bg); top.border = BRD
    mid = ws.cell(row+1, col)
    cell_style(mid, value, bold=True, sz=16, bg=bg, color=C_WHITE, align='center')
    mid.border = BRD
    bot = ws.cell(row+2, col)
    cell_style(bot, label, bold=False, sz=9, bg=bg, color='D6E4F0', align='center', wrap=True)
    bot.border = BRD
    ws.row_dimensions[row].height = 6
    ws.row_dimensions[row+1].height = 30
    ws.row_dimensions[row+2].height = 22

# ─── WORKBOOK ────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)
n = len(marketplaces)

# ════════════════════════════════════════════════════════════════
# ABA 1 — ANÁLISE BASE
# ════════════════════════════════════════════════════════════════
ws = wb.create_sheet('1. Análise Base')
title_row(ws, '📊  ANÁLISE BASE — Horas e Pessoas por Marketplace', 6)

# KPI cards (row 3-5)
ws.row_dimensions[2].height = 8
kpi_card(ws, 3, 1, 'Total de Repasses/Mês', int(total_repasses), C_CARD1)
kpi_card(ws, 3, 3, 'Total de Horas/Mês', f'{total_horas:.0f}h', C_CARD2)
kpi_card(ws, 3, 5, 'Pessoas Necessárias (total)', f'{sum(pessoas_parciais.values()):.2f}', C_CARD3)

ws.row_dimensions[6].height = 8
hdr_row(ws, ['Marketplace', 'Repasses/Mês', 'Tempo/Repasse (h)', 'Horas Totais/Mês', 'Pessoas Necessárias', '% do Esforço Total'], 7)
for i, m in enumerate(marketplaces):
    data_row(ws, 8+i, [
        m,
        int(qtd_repasses[m]),
        tempo_repasse[m],
        round(horas_totais[m], 1),
        round(pessoas_parciais[m], 3),
        f'{pct_horas[m]*100:.1f}%'
    ], alt=i%2==1)
tr = 8 + n
total_row(ws, tr, ['TOTAL', int(total_repasses), '—', round(total_horas,1), round(sum(pessoas_parciais.values()),3), '100%'])

# Gráfico barras — Horas totais
chart = BarChart()
chart.type = 'col'
chart.title = 'Horas Totais por Marketplace'
chart.y_axis.title = 'Horas/Mês'
chart.style = 10
chart.height = 12; chart.width = 18
data_ref = Reference(ws, min_col=4, min_row=7, max_row=7+n-1)
cats    = Reference(ws, min_col=1, min_row=8, max_row=7+n)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats)
chart.series[0].graphicalProperties.solidFill = '2E75B6'
ws.add_chart(chart, 'A' + str(tr + 2))
set_widths(ws, [22, 14, 18, 16, 18, 16])

# ════════════════════════════════════════════════════════════════
# ABA 2 — ÍNDICE DE COMPLEXIDADE
# ════════════════════════════════════════════════════════════════
ws = wb.create_sheet('2. Complexidade')
title_row(ws, '⚡  ÍNDICE DE COMPLEXIDADE — Quão Lento é Cada Canal', 5)
ws.row_dimensions[2].height = 8

kpi_card(ws, 3, 1, 'Canal Mais Rápido', mais_rapido, C_CARD1)
kpi_card(ws, 3, 3, 'Tempo de Referência', f'{min_tempo:.0f}h/repasse', C_CARD2)
max_idx = max(indice_complexidade, key=indice_complexidade.get)
kpi_card(ws, 3, 5, 'Canal Mais Lento', f'{max_idx} ({indice_complexidade[max_idx]:.0f}x)', C_CARD4)

ws.row_dimensions[6].height = 8
hdr_row(ws, ['Marketplace', 'Tempo/Repasse (h)', 'Índice', 'Vezes mais lento', 'Status'], 7)
for i, m in enumerate(marketplaces):
    idx = indice_complexidade[m]
    is_ref = m == mais_rapido
    bg_status = {5: C_OK if is_ref else (C_WARN if idx <= 5 else C_RED)}
    status = '★ Referência' if is_ref else (f'{idx:.1f}x mais lento' if idx <= 10 else f'⚠ {idx:.1f}x mais lento')
    data_row(ws, 8+i, [m, tempo_repasse[m], round(idx,2), '' if is_ref else f'{idx-1:.1f}x', status],
             alt=i%2==1, custom_bg=bg_status)

# Gráfico
chart = BarChart()
chart.type = 'col'
chart.title = 'Índice de Complexidade por Canal'
chart.y_axis.title = 'Índice (1 = mais rápido)'
chart.style = 10; chart.height = 12; chart.width = 18
data_ref = Reference(ws, min_col=3, min_row=7, max_row=7+n)
cats    = Reference(ws, min_col=1, min_row=8, max_row=7+n)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats)
chart.series[0].graphicalProperties.solidFill = 'C00000'
ws.add_chart(chart, 'A' + str(8+n+2))
set_widths(ws, [22, 18, 10, 16, 22])

# ════════════════════════════════════════════════════════════════
# ABA 3 — CONCENTRAÇÃO DE ESFORÇO
# ════════════════════════════════════════════════════════════════
ws = wb.create_sheet('3. Concentração Esforço')
title_row(ws, '🎯  CONCENTRAÇÃO DE ESFORÇO — Volume vs. Tempo Consumido', 6)
ws.row_dimensions[2].height = 8

ineficientes = [m for m in marketplaces if delta_esforco[m] > 0]
maior_delta  = max(marketplaces, key=lambda m: delta_esforco[m])
kpi_card(ws, 3, 1, 'Canais Ineficientes', str(len(ineficientes)), C_CARD4)
kpi_card(ws, 3, 3, 'Maior Desvio', f'{maior_delta} (+{delta_esforco[maior_delta]*100:.1f}%)', C_RED[:-2] + '99' if False else C_CARD4)
kpi_card(ws, 3, 5, 'Total Horas/Mês', f'{total_horas:.0f}h', C_CARD2)
kpi_card(ws, 3, 7, 'Total Repasses/Mês', int(total_repasses), C_CARD1)

ws.row_dimensions[6].height = 8
hdr_row(ws, ['Marketplace', '% Volume (Repasses)', '% Esforço (Horas)', 'Δ Diferença', 'Interpretação', 'Status'], 7, bg=C_MID)
for i, m in enumerate(marketplaces):
    d = delta_esforco[m]
    status = '⚠ Consome mais' if d > 0 else '✓ Eficiente'
    interp = 'Esforço desproporcional ao volume' if d > 0 else 'Abaixo da média de esforço'
    bg_s = {5: C_RED if d > 0.1 else (C_WARN if d > 0 else C_OK),
            6: C_RED if d > 0.1 else (C_WARN if d > 0 else C_OK)}
    data_row(ws, 8+i, [m, f'{pct_repasses[m]*100:.1f}%', f'{pct_horas[m]*100:.1f}%',
                        f'{d*100:+.1f}%', interp, status], alt=i%2==1, custom_bg=bg_s)

# Gráfico barras agrupadas
chart = BarChart()
chart.type = 'col'
chart.title = '% Volume vs % Esforço por Canal'
chart.y_axis.title = 'Percentual (%)'
chart.style = 10; chart.height = 12; chart.width = 22
d1 = Reference(ws, min_col=2, min_row=7, max_row=7+n)
d2 = Reference(ws, min_col=3, min_row=7, max_row=7+n)
cats = Reference(ws, min_col=1, min_row=8, max_row=7+n)
chart.add_data(d1, titles_from_data=True)
chart.add_data(d2, titles_from_data=True)
chart.set_categories(cats)
chart.grouping = 'clustered'
chart.series[0].graphicalProperties.solidFill = '2E75B6'
chart.series[1].graphicalProperties.solidFill = 'C00000'
ws.add_chart(chart, 'A' + str(8+n+2))
set_widths(ws, [22, 20, 20, 14, 30, 16])

# ════════════════════════════════════════════════════════════════
# ABA 4 — CUSTO DE OPORTUNIDADE
# ════════════════════════════════════════════════════════════════
ws = wb.create_sheet('4. Custo Oportunidade')
title_row(ws, '💡  CUSTO DE OPORTUNIDADE — Horas Perdidas por Ineficiência', 6)
ws.row_dimensions[2].height = 8

total_poupadas   = sum(horas_poupadas.values())
total_p_perdidas = sum(pessoas_perdidas.values())
custo_inefic     = total_poupadas * CUSTO_HORA
kpi_card(ws, 3, 1, 'Horas Poupadas (cenário ideal)', f'{total_poupadas:.0f}h/mês', C_CARD4)
kpi_card(ws, 3, 3, 'Pessoas Liberadas (cenário ideal)', f'{total_p_perdidas:.2f}', C_CARD2)
kpi_card(ws, 3, 5, 'Custo da Ineficiência/Mês', f'R$ {custo_inefic:,.0f}', C_CARD1)

ws.row_dimensions[6].height = 8
hdr_row(ws, ['Marketplace', 'Horas Atuais', 'Horas (Cenário Ideal)', 'Horas Poupadas', 'Pessoas Liberadas', 'Custo da Ineficiência (R$)'], 7)
for i, m in enumerate(marketplaces):
    ci = horas_poupadas[m] * CUSTO_HORA
    bg = {4: C_WARN if horas_poupadas[m] > 50 else C_OK,
          5: C_WARN if pessoas_perdidas[m] > 0.3 else C_OK}
    data_row(ws, 8+i, [m, round(horas_totais[m],1), round(horas_ideais[m],1),
                        round(horas_poupadas[m],1), round(pessoas_perdidas[m],3),
                        f'R$ {ci:,.2f}'], alt=i%2==1, custom_bg=bg)
tr = 8+n
total_row(ws, tr, ['TOTAL', round(total_horas,1), round(sum(horas_ideais.values()),1),
                    round(total_poupadas,1), round(total_p_perdidas,3), f'R$ {custo_inefic:,.2f}'])

chart = BarChart()
chart.type = 'col'
chart.title = 'Horas Poupadas por Canal (Cenário Ideal)'
chart.y_axis.title = 'Horas/Mês'
chart.style = 10; chart.height = 12; chart.width = 18
d1 = Reference(ws, min_col=4, min_row=7, max_row=7+n)
cats = Reference(ws, min_col=1, min_row=8, max_row=7+n)
chart.add_data(d1, titles_from_data=True)
chart.set_categories(cats)
chart.series[0].graphicalProperties.solidFill = 'FF6600'
ws.add_chart(chart, 'A' + str(tr+2))
set_widths(ws, [22, 14, 22, 16, 18, 24])

# ════════════════════════════════════════════════════════════════
# ABA 5 — SENSIBILIDADE DE TEMPO
# ════════════════════════════════════════════════════════════════
ws = wb.create_sheet('5. Sensibilidade Tempo')
title_row(ws, f'⏱  SENSIBILIDADE DE TEMPO — Impacto se Repasse Ficar Mais Lento ou Mais Rápido', 8)
ws.row_dimensions[2].height = 8

total_h_lento  = sum(horas_lento.values())
total_h_rapido = sum(horas_rapido.values())
total_p_lento  = sum(pessoas_lento.values())
total_p_rapido = sum(pessoas_rapido.values())

kpi_card(ws, 3, 1, 'Horas Atuais/Mês',          f'{total_horas:.0f}h',      C_CARD1)
kpi_card(ws, 3, 3, f'Horas Cenário Lento (+{VARIACAO_TEMPO_LENTO*100:.0f}%)',  f'{total_h_lento:.0f}h',   C_CARD4)
kpi_card(ws, 3, 5, f'Horas Cenário Rápido (-{VARIACAO_TEMPO_RAPIDO*100:.0f}%)', f'{total_h_rapido:.0f}h',  C_CARD3)
kpi_card(ws, 3, 7, 'Custo Extra (cenário lento)', f'R$ {sum(delta_custo_lento.values()):,.0f}', C_CARD4)

ws.row_dimensions[6].height = 8

# Subtítulo cenário lento
r_sub = 7
ws.merge_cells(start_row=r_sub, start_column=1, end_row=r_sub, end_column=8)
c_sub = ws.cell(r_sub, 1, f'  ⚠  CENÁRIO PESSIMISTA — Repasse fica {VARIACAO_TEMPO_LENTO*100:.0f}% mais lento')
c_sub.font = Font(name='Arial', bold=True, size=11, color=C_WHITE)
c_sub.fill = fill('833C00'); c_sub.alignment = Alignment(horizontal='left', vertical='center'); c_sub.border = BRD
ws.row_dimensions[r_sub].height = 22

hdr_row(ws, ['Marketplace', 'Tempo Atual (h)', f'Tempo Lento (h)', 'Horas Atuais', 'Horas Lento', 'Δ Horas', 'Pessoas Lento', 'Δ Custo (R$)'], r_sub+1, bg='833C00')
for i, m in enumerate(marketplaces):
    bg = {6: C_RED if delta_h_lento[m] > 20 else C_WARN,
          8: C_RED if delta_custo_lento[m] > 500 else C_WARN}
    data_row(ws, r_sub+2+i, [
        m,
        round(tempo_repasse[m], 1),
        round(tempo_lento[m], 1),
        round(horas_totais[m], 1),
        round(horas_lento[m], 1),
        f'+{delta_h_lento[m]:.1f}h',
        round(pessoas_lento[m], 3),
        f'+R$ {delta_custo_lento[m]:,.2f}',
    ], alt=i%2==1, custom_bg=bg)
tr_lento = r_sub+2+n
total_row(ws, tr_lento, ['TOTAL', '—', '—',
    round(total_horas,1), round(total_h_lento,1),
    f'+{sum(delta_h_lento.values()):.1f}h',
    round(total_p_lento,3),
    f'+R$ {sum(delta_custo_lento.values()):,.2f}'])

ws.row_dimensions[tr_lento+1].height = 10

# Subtítulo cenário rápido
r_sub2 = tr_lento + 2
ws.merge_cells(start_row=r_sub2, start_column=1, end_row=r_sub2, end_column=8)
c_sub2 = ws.cell(r_sub2, 1, f'  ✅  CENÁRIO OTIMISTA — Repasse fica {VARIACAO_TEMPO_RAPIDO*100:.0f}% mais rápido')
c_sub2.font = Font(name='Arial', bold=True, size=11, color=C_WHITE)
c_sub2.fill = fill('375623'); c_sub2.alignment = Alignment(horizontal='left', vertical='center'); c_sub2.border = BRD
ws.row_dimensions[r_sub2].height = 22

hdr_row(ws, ['Marketplace', 'Tempo Atual (h)', f'Tempo Rápido (h)', 'Horas Atuais', 'Horas Rápido', 'Δ Horas', 'Pessoas Rápido', 'Δ Custo (R$)'], r_sub2+1, bg='375623')
for i, m in enumerate(marketplaces):
    data_row(ws, r_sub2+2+i, [
        m,
        round(tempo_repasse[m], 1),
        round(tempo_rapido[m], 1),
        round(horas_totais[m], 1),
        round(horas_rapido[m], 1),
        f'{delta_h_rapido[m]:.1f}h',
        round(pessoas_rapido[m], 3),
        f'R$ {delta_custo_rapido[m]:,.2f}',
    ], alt=i%2==1, custom_bg={6: C_OK, 8: C_OK})
tr_rapido = r_sub2+2+n
total_row(ws, tr_rapido, ['TOTAL', '—', '—',
    round(total_horas,1), round(total_h_rapido,1),
    f'{sum(delta_h_rapido.values()):.1f}h',
    round(total_p_rapido,3),
    f'R$ {sum(delta_custo_rapido.values()):,.2f}'])

# Gráfico comparativo 3 cenários
chart = BarChart()
chart.type = 'col'
chart.title = 'Comparativo de Horas: Atual vs. Lento vs. Rápido'
chart.y_axis.title = 'Horas/Mês'
chart.grouping = 'clustered'
chart.style = 10; chart.height = 14; chart.width = 22

# Usar dados do bloco lento para o gráfico (contém atual e lento)
d_atual = Reference(ws, min_col=4, min_row=r_sub+1, max_row=r_sub+1+n)
d_lento = Reference(ws, min_col=5, min_row=r_sub+1, max_row=r_sub+1+n)
d_rapid = Reference(ws, min_col=5, min_row=r_sub2+1, max_row=r_sub2+1+n)
cats    = Reference(ws, min_col=1, min_row=r_sub+2, max_row=r_sub+1+n)
chart.add_data(d_atual, titles_from_data=True)
chart.add_data(d_lento, titles_from_data=True)
chart.add_data(d_rapid, titles_from_data=True)
chart.set_categories(cats)
chart.series[0].graphicalProperties.solidFill = '2E75B6'
chart.series[1].graphicalProperties.solidFill = 'C00000'
chart.series[2].graphicalProperties.solidFill = '375623'
ws.add_chart(chart, 'A' + str(tr_rapido + 2))

set_widths(ws, [22, 16, 16, 14, 14, 12, 14, 20])

# ════════════════════════════════════════════════════════════════
# ABA 6 — PAYBACK DE AUTOMAÇÃO
# ════════════════════════════════════════════════════════════════
ws = wb.create_sheet('6. Payback Automação')
title_row(ws, f'🤖  PAYBACK DE AUTOMAÇÃO — Investimento de R$ {CUSTO_AUTOMACAO:,.0f}', 5)
ws.row_dimensions[2].height = 8

melhor_pb = min(marketplaces, key=lambda m: payback_meses[m])
kpi_card(ws, 3, 1, 'Custo Mensal Total (todos)', f'R$ {sum(custo_mensal.values()):,.0f}', C_CARD1)
kpi_card(ws, 3, 3, 'Menor Payback', f'{melhor_pb} ({payback_meses[melhor_pb]:.1f} meses)', C_CARD3)
kpi_card(ws, 3, 5, 'Investimento Estimado', f'R$ {CUSTO_AUTOMACAO:,.0f}', C_CARD2)

ws.row_dimensions[6].height = 8
hdr_row(ws, ['Marketplace', 'Horas/Mês', 'Custo Mensal (R$)', 'Custo Automação (R$)', 'Payback (meses)', 'Recomendação'], 7)
for i, m in enumerate(marketplaces):
    pb = payback_meses[m]
    rec = '✓ Excelente ROI' if pb <= 6 else ('➜ Bom ROI' if pb <= 12 else '⚠ ROI longo')
    bg = {5: C_OK if pb <= 6 else (C_WARN if pb <= 12 else C_RED),
          6: C_OK if pb <= 6 else (C_WARN if pb <= 12 else C_RED)}
    data_row(ws, 8+i, [m, round(horas_totais[m],1), f'R$ {custo_mensal[m]:,.2f}',
                        f'R$ {CUSTO_AUTOMACAO:,.2f}', round(pb,1), rec],
             alt=i%2==1, custom_bg=bg)

chart = BarChart()
chart.type = 'bar'
chart.title = 'Payback por Canal (meses)'
chart.x_axis.title = 'Meses para retorno'
chart.style = 10; chart.height = 12; chart.width = 18
d1 = Reference(ws, min_col=5, min_row=7, max_row=7+n)
cats = Reference(ws, min_col=1, min_row=8, max_row=7+n)
chart.add_data(d1, titles_from_data=True)
chart.set_categories(cats)
chart.series[0].graphicalProperties.solidFill = '375623'
ws.add_chart(chart, 'A' + str(8+n+2))
set_widths(ws, [22, 12, 20, 20, 16, 20])

# ════════════════════════════════════════════════════════════════
# ABA 7 — RANKING DE PRIORIDADE
# ════════════════════════════════════════════════════════════════
ws = wb.create_sheet('7. Ranking Prioridade')
title_row(ws, '🏆  RANKING DE PRIORIDADE — Canais para Atacar Primeiro', 5)
ws.row_dimensions[2].height = 8

kpi_card(ws, 3, 1, '1º Prioridade', ranking[0], C_CARD4)
kpi_card(ws, 3, 3, '2º Prioridade', ranking[1] if len(ranking) > 1 else '—', C_CARD2)
kpi_card(ws, 3, 5, 'Custo Top-2 / Mês', f'R$ {custo_mensal[ranking[0]]+custo_mensal[ranking[1]]:,.0f}', C_CARD1)

ws.row_dimensions[6].height = 8
hdr_row(ws, ['Posição', 'Marketplace', 'Score de Prioridade', 'Custo Mensal (R$)', 'Payback (meses)', 'Ação Recomendada'], 7)
medals = {0: 'FFD700', 1: 'C0C0C0', 2: 'CD7F32'}
for i, m in enumerate(ranking):
    pb = payback_meses[m]
    acao = 'Automatizar imediatamente' if i < 2 else ('Planejar automação' if i < 4 else 'Monitorar')
    bg = {}
    if i in medals:
        bg = {1: medals[i], 2: medals[i]}
    data_row(ws, 8+i, [i+1, m, round(score_ranking[m],2), f'R$ {custo_mensal[m]:,.2f}',
                        round(pb,1), acao], custom_bg=bg)

chart = BarChart()
chart.type = 'bar'
chart.title = 'Score de Prioridade por Canal'
chart.x_axis.title = 'Score'
chart.style = 10; chart.height = 12; chart.width = 18
d1 = Reference(ws, min_col=3, min_row=7, max_row=7+n)
cats = Reference(ws, min_col=2, min_row=8, max_row=7+n)
chart.add_data(d1, titles_from_data=True)
chart.set_categories(cats)
chart.series[0].graphicalProperties.solidFill = '4B0082'
ws.add_chart(chart, 'A' + str(8+n+2))
set_widths(ws, [10, 22, 20, 20, 16, 24])

# ════════════════════════════════════════════════════════════════
# ABA DASHBOARD — PAINEL EXECUTIVO
# ════════════════════════════════════════════════════════════════
ws = wb.create_sheet('📊 Dashboard', 0)
title_row(ws, '📊  PAINEL EXECUTIVO — CONCILIAÇÃO DE MARKETPLACE', 8, bg=C_DARK, sz=14)
ws.row_dimensions[1].height = 36
ws.row_dimensions[2].height = 10

# Bloco de KPIs principais — linha 3
kpi_card(ws, 3, 1, 'Total Repasses/Mês',    int(total_repasses),              C_CARD1)
kpi_card(ws, 3, 3, 'Total Horas/Mês',       f'{total_horas:.0f}h',            C_CARD2)
kpi_card(ws, 3, 5, 'Pessoas Necessárias',   f'{sum(pessoas_parciais.values()):.2f}', C_CARD3)
kpi_card(ws, 3, 7, 'Custo Operacional/Mês', f'R$ {sum(custo_mensal.values()):,.0f}', C_CARD4)

ws.row_dimensions[6].height = 10

# Bloco 2 — linha 7
kpi_card(ws, 7, 1, 'Horas Poupadas (ideal)', f'{sum(horas_poupadas.values()):.0f}h', C_CARD5)
kpi_card(ws, 7, 3, 'Custo da Ineficiência', f'R$ {sum(horas_poupadas.values())*CUSTO_HORA:,.0f}', C_CARD4)
kpi_card(ws, 7, 5, f'Horas se Tudo Desacelerar +{VARIACAO_TEMPO_LENTO*100:.0f}%', f'{sum(horas_lento.values()):.0f}h', C_CARD4)
kpi_card(ws, 7, 7, f'Horas se Tudo Acelerar -{VARIACAO_TEMPO_RAPIDO*100:.0f}%', f'{sum(horas_rapido.values()):.0f}h', C_CARD3)

ws.row_dimensions[10].height = 10

# Seção: Tabela resumo consolidada
r = 11
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
c = ws.cell(r, 1, '  VISÃO CONSOLIDADA POR MARKETPLACE')
c.font = Font(name='Arial', bold=True, size=11, color=C_WHITE)
c.fill = fill(C_MID); c.alignment = Alignment(horizontal='left', vertical='center'); c.border = BRD
ws.row_dimensions[r].height = 24

hdr_row(ws, ['Marketplace', 'Repasses', 'Horas/Mês', 'Pessoas', 'Índice Compl.', '% Esforço', 'Custo/Mês (R$)', 'Payback (m)'], r+1, bg=C_LIGHT)
for i, m in enumerate(marketplaces):
    pb = payback_meses[m]
    bg_pb  = {8: C_OK if pb <= 6 else (C_WARN if pb <= 12 else C_RED)}
    bg_idx = {5: C_RED if indice_complexidade[m] > 10 else (C_WARN if indice_complexidade[m] > 3 else C_OK)}
    bg_all = {**bg_pb, **bg_idx}
    data_row(ws, r+2+i, [
        m,
        int(qtd_repasses[m]),
        round(horas_totais[m],1),
        round(pessoas_parciais[m],3),
        round(indice_complexidade[m],1),
        f'{pct_horas[m]*100:.1f}%',
        f'R$ {custo_mensal[m]:,.0f}',
        round(pb,1)
    ], alt=i%2==1, custom_bg=bg_all)
tr_dash = r+2+n
total_row(ws, tr_dash, ['TOTAL', int(total_repasses), round(total_horas,1),
                         round(sum(pessoas_parciais.values()),3), '—', '100%',
                         f'R$ {sum(custo_mensal.values()):,.0f}', '—'])

ws.row_dimensions[tr_dash+1].height = 10

# Seção: Ranking
rr = tr_dash + 2
ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=8)
c2 = ws.cell(rr, 1, '  🏆  RANKING DE PRIORIDADE PARA AUTOMAÇÃO')
c2.font = Font(name='Arial', bold=True, size=11, color=C_WHITE)
c2.fill = fill(C_DARK); c2.alignment = Alignment(horizontal='left', vertical='center'); c2.border = BRD
ws.row_dimensions[rr].height = 24

hdr_row(ws, ['Pos.', 'Marketplace', 'Score', 'Custo/Mês (R$)', 'Payback (m)', 'Ação', '', ''], rr+1, bg=C_MID)
for i, m in enumerate(ranking):
    pb  = payback_meses[m]
    acao = 'Automatizar imediatamente' if i < 2 else ('Planejar automação' if i < 4 else 'Monitorar')
    bg  = {}
    if i == 0: bg = {1: 'FFD700', 2: 'FFD700'}
    elif i == 1: bg = {1: 'C0C0C0', 2: 'C0C0C0'}
    elif i == 2: bg = {1: 'CD7F32', 2: 'CD7F32'}
    data_row(ws, rr+2+i, [i+1, m, round(score_ranking[m],0),
                            f'R$ {custo_mensal[m]:,.0f}', round(pb,1), acao, '', ''],
             custom_bg=bg)

# Gráfico pizza — distribuição de horas
pie = PieChart()
pie.title = 'Distribuição de Horas por Canal'
pie.style = 10; pie.height = 12; pie.width = 16
labels = Reference(ws, min_col=2, min_row=r+2, max_row=r+1+n)
data   = Reference(ws, min_col=3, min_row=r+1, max_row=r+1+n)
pie.add_data(data, titles_from_data=True)
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
ws.add_chart(pie, get_column_letter(1) + str(rr+2+n+2))

set_widths(ws, [22, 10, 12, 12, 14, 12, 18, 12])

# ─── SALVAR ──────────────────────────────────────────────────────
out_path = os.path.join(BASE_DIR, 'analise_conciliacao_marketplace.xlsx')
wb.save(out_path)
print(f'✅ Excel gerado: {out_path}')
